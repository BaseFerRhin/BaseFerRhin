"""CSV / Excel row extraction with encoding and delimiter detection."""

from __future__ import annotations

import csv
import logging
from dataclasses import fields
from io import StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.domain.models.raw_record import RawRecord

logger = logging.getLogger(__name__)

_ENCODINGS: tuple[str, ...] = ("utf-8", "latin-1", "cp1252")
_DELIMS: tuple[str, ...] = (",", ";", "\t")
_RECORD_FIELDS = {f.name for f in fields(RawRecord)} - {"extra"}


class CSVExtractor:
    """Maps **source** column headers to ``RawRecord`` attribute names."""

    def __init__(self, column_mapping: dict[str, str] | None = None) -> None:
        # keys: CSV/Excel column name; values: RawRecord field (e.g. ``type`` → ``type_mention``).
        self._column_mapping: dict[str, str] = dict(column_mapping or {})

    def supported_formats(self) -> list[str]:
        return [".csv", ".xlsx"]

    def extract(self, source_path: Path) -> list[RawRecord]:
        path_str = str(source_path.resolve())
        suffix = source_path.suffix.lower()
        if suffix == ".xlsx":
            return self._extract_xlsx(source_path, path_str)
        return self._extract_csv(source_path, path_str)

    def _extract_xlsx(self, source_path: Path, path_str: str) -> list[RawRecord]:
        wb = load_workbook(source_path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            wb.close()
            return []
        names = [str(c or "").strip() for c in header]
        out: list[RawRecord] = []
        for row in rows_iter:
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue
            row_dict = {names[j]: row[j] for j in range(min(len(names), len(row)))}
            out.append(self._row_to_record(row_dict, path_str, "xlsx"))
        wb.close()
        logger.info("xlsx extract: %s rows=%d", source_path.name, len(out))
        return out

    def _extract_csv(self, source_path: Path, path_str: str) -> list[RawRecord]:
        raw_bytes = source_path.read_bytes()
        text, delim = self._decode_and_delim(raw_bytes)
        reader = csv.DictReader(StringIO(text), delimiter=delim)
        out = [self._row_to_record(dict(r), path_str, "csv") for r in reader if any(r.values())]
        logger.info("csv extract: %s rows=%d", source_path.name, len(out))
        return out

    def _decode_and_delim(self, raw_bytes: bytes) -> tuple[str, str]:
        last_err: Exception | None = None
        for enc in _ENCODINGS:
            try:
                text = raw_bytes.decode(enc)
            except UnicodeDecodeError as e:
                last_err = e
                continue
            delim = self._sniff_delim(text)
            return text, delim
        raise ValueError(f"Could not decode CSV with encodings {_ENCODINGS}") from last_err

    def _sniff_delim(self, sample: str) -> str:
        head = "\n".join(sample.splitlines()[:5])
        try:
            dialect = csv.Sniffer().sniff(head, delimiters="".join(_DELIMS))
            if dialect.delimiter in _DELIMS:
                return dialect.delimiter
        except csv.Error:
            pass
        counts = {d: head.count(d) for d in _DELIMS}
        return max(counts, key=counts.get)

    def _row_to_record(self, row: dict[str, Any], path_str: str, method: str) -> RawRecord:
        kwargs: dict[str, Any] = {
            "source_path": path_str,
            "extraction_method": method,
            "extra": {},
        }
        for src_col, val in row.items():
            key = str(src_col).strip()
            target = self._column_mapping.get(key, key)
            if target in _RECORD_FIELDS:
                kwargs[target] = self._coerce_field(target, val)
            else:
                kwargs["extra"][key] = val
        if not kwargs.get("raw_text"):
            kwargs["raw_text"] = " | ".join(
                f"{k}={v}" for k, v in row.items() if v not in (None, "")
            )
        return RawRecord(**kwargs)

    def _coerce_field(self, field: str, val: Any) -> Any:
        if val is None or (isinstance(val, str) and val.strip() == ""):
            return None if field != "raw_text" else ""
        if field in ("latitude_raw", "longitude_raw"):
            try:
                return float(str(val).replace(",", "."))
            except ValueError:
                return None
        if field == "page_number":
            try:
                return int(val)
            except (TypeError, ValueError):
                return None
        if field == "raw_text":
            return str(val).strip()
        return str(val).strip() if not isinstance(val, str) else val.strip()
