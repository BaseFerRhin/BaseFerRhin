#!/usr/bin/env python3
"""Ingestion T1–T6: Alsace_Basel_AF (1).xlsx → sites_cleaned.csv + quality_report.json."""

from __future__ import annotations

import json
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from pyproj import Transformer

REPO_ROOT = Path(__file__).resolve().parents[3]
ANALYSIS_DIR = Path(__file__).resolve().parent

SOURCE_CANDIDATES = [
    REPO_ROOT / "data/input/Alsace_Basel_AF (1).xlsx",
    REPO_ROOT / "RawData/GrosFichiers - Béhague/Alsace_Basel_AF (1).xlsx",
]

EXPECTED_COLS = [
    "id_site", "pays", "admin1", "commune", "lieu_dit", "lieu_dit_autre",
    "x", "y", "epsg_coord", "decouverte_annee", "decouverte_operation",
    "ref_biblio", "ref_rapport", "auteur", "date", "commentaire",
]

WGS84_LON_RANGE = (5.0, 11.0)
WGS84_LAT_RANGE = (45.0, 50.0)
DEDUP_METERS = 500.0
FUZZY_THRESHOLD = 85

# Type keys → normalized (archéo-proto-eu)
TYPE_NORMALIZE = {
    "OPPIDUM": "OPPIDUM",
    "HABITAT": "HABITAT",
    "NECROPOLE": "NECROPOLE",
    "DEPOT": "DEPOT",
    "SANCTUAIRE": "SANCTUAIRE",
    "ATELIER": "INDETERMINE",
    "VOIE": "INDETERMINE",
    "TUMULUS": "NECROPOLE",
}


def _patch_openpyxl_load():
    import openpyxl.worksheet.datavalidation as dv_mod

    original_init = dv_mod.DataValidation.__init__

    def _patched_init(self_dv, *args, **kwargs):
        sqref = kwargs.get("sqref") or (args[6] if len(args) > 6 else None)
        if sqref is not None and not isinstance(sqref, str):
            kwargs["sqref"] = str(sqref)
            if len(args) > 6:
                args = list(args)
                args[6] = str(sqref)
                args = tuple(args)
        try:
            original_init(self_dv, *args, **kwargs)
        except TypeError:
            kwargs.pop("sqref", None)
            original_init(self_dv, *args, **kwargs)

    dv_mod.DataValidation.__init__ = _patched_init
    try:
        from openpyxl import load_workbook

        return load_workbook, dv_mod, original_init
    except Exception:
        dv_mod.DataValidation.__init__ = original_init
        raise


def load_sites_sheet(xlsx_path: Path) -> pd.DataFrame:
    load_workbook, dv_mod, original_init = _patch_openpyxl_load()
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    finally:
        dv_mod.DataValidation.__init__ = original_init
    if "sites" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"No 'sites' sheet in {xlsx_path}")
    ws = wb["sites"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    header = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
    df = pd.DataFrame(rows[1:], columns=header)
    df = df.loc[:, [c for c in df.columns if c and not str(c).startswith("col_") or df[c].notna().any()]]
    # Drop trailing unnamed empty columns
    df = df.dropna(axis=1, how="all")
    rename_map = {c: c for c in df.columns}
    for c in list(df.columns):
        if str(c).lower().startswith("unnamed"):
            df = df.drop(columns=[c], errors="ignore")
    return df


def parse_float(val: Any) -> float | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).replace("\xa0", " ").strip().replace(",", ".")
    if not s or s.lower() == "nan":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_epsg(val: Any) -> int | None:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        return int(round(float(val)))
    except (ValueError, TypeError):
        return None


def haversine_m(lon1: float, lat1: float, lon2: float, lat2: float) -> float:
    r = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(min(1.0, math.sqrt(a)))


def load_toponyme_map(path: Path) -> dict[str, set[str]]:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    variant_to_canon: dict[str, set[str]] = defaultdict(set)
    for entry in data.get("concordance", []):
        canon = str(entry.get("canonical", "")).strip().lower()
        if not canon:
            continue
        variant_to_canon[canon].add(canon)
        for v in entry.get("variants", []):
            vv = str(v).strip().lower()
            if vv:
                variant_to_canon[vv].add(canon)
    return dict(variant_to_canon)


def commune_tokens(norm: str, toponyme_map: dict[str, set[str]]) -> set[str]:
    n = norm.strip().lower()
    out = {n}
    for v, canons in toponyme_map.items():
        if v == n or v in n or n in v:
            out |= canons
    return out


def load_period_patterns(periodes_path: Path) -> list[tuple[str, str, dict]]:
    """(match_key, pattern, meta with date_debut, date_fin, sous)."""
    with open(periodes_path, encoding="utf-8") as f:
        data = json.load(f)
    periodes = data.get("periodes", {})
    items: list[tuple[str, str, dict]] = []
    display = {"HALLSTATT": "Hallstatt", "LA_TENE": "La Tène", "TRANSITION": "Transition"}
    for key, block in periodes.items():
        meta_base = {
            "periode": display.get(key, key),
            "date_debut": block.get("date_debut"),
            "date_fin": block.get("date_fin"),
        }
        for pat in block.get("patterns_fr", []) + block.get("patterns_de", []):
            if pat:
                items.append((key, pat, dict(meta_base)))
        for sous_name, sous in (block.get("sous_periodes") or {}).items():
            sm = dict(meta_base)
            sm["sous_periode"] = sous_name
            sm["date_debut"] = sous.get("date_debut", sm["date_debut"])
            sm["date_fin"] = sous.get("date_fin", sm["date_fin"])
            items.append((key, sous_name, sm))
    items.sort(key=lambda x: len(x[1]), reverse=True)
    return items


def load_type_patterns(types_path: Path) -> list[tuple[str, str]]:
    with open(types_path, encoding="utf-8") as f:
        data = json.load(f)
    pairs: list[tuple[str, str]] = []
    for key, langs in (data.get("aliases") or {}).items():
        for lang in ("fr", "de"):
            for alias in langs.get(lang, []) or []:
                a = str(alias).strip()
                if a:
                    pairs.append((key, a))
    pairs.sort(key=lambda x: len(x[1]), reverse=True)
    return pairs


def infer_period(text: str, patterns: list[tuple[str, str, dict]]) -> tuple[dict | None, str | None]:
    if not text:
        return None, None
    t = text.lower()
    for _key, pat, meta in patterns:
        if pat.lower() in t:
            return meta, pat
    return None, None


def infer_type(text: str, type_pairs: list[tuple[str, str]]) -> tuple[str | None, str | None]:
    if not text:
        return None, None
    t = text.lower()
    for key, alias in type_pairs:
        if alias.lower() in t:
            return TYPE_NORMALIZE.get(key, "INDETERMINE"), alias
    return None, None


def row_xy_epsg(row: pd.Series) -> tuple[float | None, float | None, int | None]:
    return parse_float(row.get("x")), parse_float(row.get("y")), parse_epsg(row.get("epsg_coord"))


def coords_key(x: float | None, y: float | None, epsg: int | None) -> tuple | None:
    if x is None or y is None or epsg is None:
        return None
    return (round(x, 6), round(y, 6), epsg)


def resolve_group_coords(
    sub: pd.DataFrame,
) -> tuple[float | None, float | None, int | None, list[str]]:
    """Pick (x,y,epsg); log conflicts."""
    notes: list[str] = []
    keys: list[tuple[float, float, int] | None] = []
    for _, r in sub.iterrows():
        x, y, e = row_xy_epsg(r)
        keys.append(coords_key(x, y, e))
    valid = [k for k in keys if k is not None]
    if not valid:
        return None, None, None, notes
    uniq = set(valid)
    if len(uniq) == 1:
        k = list(uniq)[0]
        return k[0], k[1], k[2], notes
    notes.append("coordinate_conflict")
    # Prefer row with valid epsg + plausible coords; else latest date
    def score_row(r: pd.Series) -> tuple:
        x, y, e = row_xy_epsg(r)
        s = 0
        if e in (4326, 25832):
            s += 10
        if e == 4326 and x is not None and y is not None:
            if WGS84_LON_RANGE[0] <= x <= WGS84_LON_RANGE[1] and WGS84_LAT_RANGE[0] <= y <= WGS84_LAT_RANGE[1]:
                s += 5
        if e == 25832 and x is not None and y is not None and x > 1000 and y > 1000:
            s += 5
        d = r.get("date")
        try:
            ts = pd.Timestamp(d).value if pd.notna(d) else 0
        except Exception:
            ts = 0
        return (s, ts)

    best = max(sub.iterrows(), key=lambda ir: score_row(ir[1]))[1]
    x, y, e = row_xy_epsg(best)
    return x, y, e, notes


def reproject_row(
    x: float, y: float, epsg: int,
    t_4326_to_2154: Transformer,
    t_25832_to_2154: Transformer,
    t_25832_to_4326: Transformer,
) -> tuple[float | None, float | None, float | None, float | None, str | None]:
    """Returns lon, lat, x_l93, y_l93, issue."""
    if epsg == 4326:
        lon, lat = x, y
        if not (WGS84_LON_RANGE[0] <= lon <= WGS84_LON_RANGE[1] and WGS84_LAT_RANGE[0] <= lat <= WGS84_LAT_RANGE[1]):
            return lon, lat, None, None, "coordinates_suspicious"
        x_l93, y_l93 = t_4326_to_2154.transform(lon, lat)
        return lon, lat, x_l93, y_l93, None
    if epsg == 25832:
        x_l93, y_l93 = t_25832_to_2154.transform(x, y)
        lon, lat = t_25832_to_4326.transform(x, y)
        return lon, lat, x_l93, y_l93, None
    return None, None, None, None, "unknown_epsg"


def merge_unique_text(series: pd.Series) -> str:
    seen: list[str] = []
    for v in series:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip()
        if s and s.lower() != "nan" and s not in seen:
            seen.append(s)
    return " | ".join(seen)


def main() -> None:
    report: dict[str, Any] = {
        "row_count_raw": 0,
        "site_count_aggregated": 0,
        "epsg_missing": [],
        "coordinates_suspicious": [],
        "coordinate_conflicts": [],
        "duplicates_with_golden_or_sites_csv": [],
        "inferred_from_text": {"period": 0, "type": 0},
        "notes": [
            "EPSG 25832: coordonnées métriques UTM, pas des degrés.",
            "Colonne date = métadonnée fichier, non utilisée comme chronologie Fer.",
        ],
        "mapping_decisions": [],
    }

    xlsx_path = next((p for p in SOURCE_CANDIDATES if p.exists()), None)
    if xlsx_path is None:
        raise FileNotFoundError("Alsace_Basel_AF (1).xlsx introuvable (data/input ou RawData).")

    df = load_sites_sheet(xlsx_path)
    report["row_count_raw"] = len(df)

    for c in EXPECTED_COLS:
        if c not in df.columns:
            raise ValueError(f"Colonne manquante: {c}")

    # T1 clean
    def strip_cell(v: Any) -> str:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        return str(v).strip()

    for col in (
        "commune", "lieu_dit", "pays", "admin1", "ref_biblio", "ref_rapport",
        "commentaire", "lieu_dit_autre", "decouverte_operation", "auteur",
    ):
        if col in df.columns:
            df[col] = df[col].map(strip_cell)

    df["_x_f"] = df["x"].map(parse_float)
    df["_y_f"] = df["y"].map(parse_float)
    df["_epsg"] = df["epsg_coord"].map(parse_epsg)

    toponyme_path = REPO_ROOT / "data/reference/toponymes_fr_de.json"
    toponyme_map = load_toponyme_map(toponyme_path) if toponyme_path.exists() else {}

    period_patterns = load_period_patterns(REPO_ROOT / "data/reference/periodes.json")
    type_patterns = load_type_patterns(REPO_ROOT / "data/reference/types_sites.json")

    t_4326_2154 = Transformer.from_crs("EPSG:4326", "EPSG:2154", always_xy=True)
    t_25832_2154 = Transformer.from_crs("EPSG:25832", "EPSG:2154", always_xy=True)
    t_25832_4326 = Transformer.from_crs("EPSG:25832", "EPSG:4326", always_xy=True)

    grouped = df.groupby("id_site", sort=True)
    report["site_count_aggregated"] = grouped.ngroups

    rows_out: list[dict[str, Any]] = []

    for id_site, sub in grouped:
        id_site = int(id_site) if not isinstance(id_site, int) else id_site
        r0 = sub.iloc[0]
        pays = str(r0.get("pays") or "").strip()
        admin1 = str(r0.get("admin1") or "").strip()
        commune = str(r0.get("commune") or "").strip()
        lieu_series = sub["lieu_dit"] if "lieu_dit" in sub.columns else pd.Series(dtype=object)
        lieu_dit = next((str(v).strip() for v in lieu_series if str(v).strip()), "")
        lieu_autre = merge_unique_text(sub["lieu_dit_autre"]) if "lieu_dit_autre" in sub.columns else ""

        bib = merge_unique_text(sub["ref_biblio"]) if "ref_biblio" in sub.columns else ""
        rap = merge_unique_text(sub["ref_rapport"]) if "ref_rapport" in sub.columns else ""
        commentaire = merge_unique_text(sub["commentaire"]) if "commentaire" in sub.columns else ""

        x, y, epsg, cnotes = resolve_group_coords(sub)
        if "coordinate_conflict" in cnotes:
            report["coordinate_conflicts"].append({"id_site": id_site, "detail": "plusieurs (x,y,epsg) dans le groupe"})

        lon, lat, x_l93, y_l93, geo_issue = (None, None, None, None, None)
        if x is not None and y is not None and epsg is not None:
            lon, lat, x_l93, y_l93, geo_issue = reproject_row(
                x, y, epsg, t_4326_2154, t_25832_2154, t_25832_4326
            )
            if geo_issue == "coordinates_suspicious":
                report["coordinates_suspicious"].append({"id_site": id_site, "x": x, "y": y, "epsg": epsg})
        text_for_infer = f"{commentaire} {lieu_dit} {lieu_autre}"
        pmeta, pmatched = infer_period(text_for_infer, period_patterns)
        tnorm, tmatched = infer_type(text_for_infer, type_patterns)

        periode = ""
        sous_periode = ""
        ddeb, dfin = "", ""
        confiance = "HIGH"
        if pmeta:
            periode = str(pmeta.get("periode") or "")
            sous_periode = str(pmeta.get("sous_periode") or "")
            ddeb = pmeta.get("date_debut")
            dfin = pmeta.get("date_fin")
            report["inferred_from_text"]["period"] += 1
            confiance = "MEDIUM"
            report["mapping_decisions"].append(
                {"id_site": id_site, "field": "periode", "rule": "pattern_text", "matched": pmatched}
            )
        type_site = tnorm or "INDETERMINE"
        if tnorm:
            report["inferred_from_text"]["type"] += 1
            confiance = "MEDIUM"
            report["mapping_decisions"].append(
                {"id_site": id_site, "field": "type_site", "rule": "types_sites.json", "matched": tmatched}
            )
        if geo_issue == "coordinates_suspicious":
            confiance = "LOW"
        if x is None or y is None:
            confiance = "LOW" if confiance == "HIGH" else confiance

        nom_site = lieu_dit or commune or f"site_{id_site}"
        dec_annee = merge_unique_text(sub["decouverte_annee"]) if "decouverte_annee" in sub.columns else ""
        dec_op = merge_unique_text(sub["decouverte_operation"]) if "decouverte_operation" in sub.columns else ""

        rows_out.append({
            "site_id": f"ALSACE-BASEL-AF-{id_site}",
            "id_site_source": id_site,
            "nom_site": nom_site,
            "commune": commune,
            "lieu_dit": lieu_dit,
            "pays": pays,
            "admin1": admin1,
            "longitude": lon,
            "latitude": lat,
            "x_l93": x_l93,
            "y_l93": y_l93,
            "epsg_source": epsg,
            "decouverte_annee": dec_annee,
            "decouverte_operation": dec_op,
            "periode": periode,
            "sous_periode": sous_periode,
            "datation_debut": ddeb if ddeb != "" else "",
            "datation_fin": dfin if dfin != "" else "",
            "type_site": type_site,
            "confiance": confiance,
            "source": "Alsace_Basel_AF_xlsx",
            "bibliographie": bib,
            "rapports": rap,
            "commentaire": commentaire,
        })

    out_df = pd.DataFrame(rows_out)
    report["epsg_missing"] = sorted(
        {int(r["id_site_source"]) for r in rows_out if r.get("epsg_source") is None}
    )
    report["sites_without_coordinates"] = sorted(
        {int(r["id_site_source"]) for r in rows_out if r.get("longitude") is None}
    )

    # T5 dedup
    try:
        from rapidfuzz import fuzz
    except ImportError:
        fuzz = None  # type: ignore

    golden_path = REPO_ROOT / "data/sources/golden_sites.csv"
    sites_csv_path = REPO_ROOT / "data/output/sites.csv"
    golden_rows: list[dict] = []
    if golden_path.exists():
        gdf = pd.read_csv(golden_path, sep=";")
        for _, gr in gdf.iterrows():
            try:
                la = float(gr["latitude_raw"])
                lo = float(gr["longitude_raw"])
            except Exception:
                continue
            golden_rows.append({
                "commune": str(gr.get("commune") or "").strip(),
                "lat": la,
                "lon": lo,
                "raw_text": str(gr.get("raw_text") or ""),
            })

    sites_refs: list[dict] = []
    if sites_csv_path.exists():
        sdf = pd.read_csv(sites_csv_path, low_memory=False)
        for _, sr in sdf.iterrows():
            try:
                xl = float(sr["x_l93"])
                yl = float(sr["y_l93"])
            except Exception:
                continue
            sites_refs.append({
                "site_id": str(sr.get("site_id", "")),
                "commune": str(sr.get("commune") or "").strip(),
                "nom_site": str(sr.get("nom_site") or "").strip(),
                "x_l93": xl,
                "y_l93": yl,
            })

    def norm_commune(c: str) -> str:
        return c.strip().lower()

    def fuzzy_ok(a: str, b: str) -> bool:
        if not fuzz or not a.strip() or not b.strip():
            return False
        return fuzz.token_sort_ratio(a.strip(), b.strip()) >= FUZZY_THRESHOLD

    for rec in rows_out:
        cid = rec["id_site_source"]
        commune_n = norm_commune(rec["commune"])
        toks = commune_tokens(commune_n, toponyme_map)
        lon, lat = rec["longitude"], rec["latitude"]
        xl93, yl93 = rec["x_l93"], rec["y_l93"]
        q_alc = f"{rec['commune']} {rec['lieu_dit']}".strip()

        for gr in golden_rows:
            g_toks = commune_tokens(norm_commune(gr["commune"]), toponyme_map)
            commune_ok = bool(toks & g_toks) or (
                norm_commune(gr["commune"]) in commune_n or commune_n in norm_commune(gr["commune"])
            )
            gq = f"{gr['commune']} {gr['raw_text'][:120]}"
            fuzzy_hit = fuzzy_ok(q_alc, gq)
            dist_ok = False
            dist_m = None
            if lon is not None and lat is not None:
                dist_m = haversine_m(lon, lat, gr["lon"], gr["lat"])
                dist_ok = dist_m < DEDUP_METERS
            if (dist_ok and commune_ok) or fuzzy_hit:
                report["duplicates_with_golden_or_sites_csv"].append({
                    "alsace_id_site": cid,
                    "match": "golden_sites",
                    "ref_commune": gr["commune"],
                    "distance_m": round(dist_m, 1) if dist_m is not None else None,
                    "via": "distance+commune" if dist_ok and commune_ok else "fuzzy",
                })
                break

        else:
            for sr in sites_refs:
                s_toks = commune_tokens(norm_commune(sr["commune"]), toponyme_map)
                commune_ok = bool(toks & s_toks) or (
                    norm_commune(sr["commune"]) in commune_n or commune_n in norm_commune(sr["commune"])
                )
                gq = f"{sr['commune']} {sr['nom_site']}"
                fuzzy_hit = fuzzy_ok(q_alc, gq)
                dist_ok = False
                dist_m = None
                if xl93 is not None and yl93 is not None:
                    dx = xl93 - sr["x_l93"]
                    dy = yl93 - sr["y_l93"]
                    dist_m = math.sqrt(dx * dx + dy * dy)
                    dist_ok = dist_m < DEDUP_METERS
                if (dist_ok and commune_ok) or fuzzy_hit:
                    report["duplicates_with_golden_or_sites_csv"].append({
                        "alsace_id_site": cid,
                        "match": "sites_csv",
                        "ref_site_id": sr["site_id"],
                        "distance_m": round(dist_m, 1) if dist_m is not None else None,
                        "via": "distance+commune" if dist_ok and commune_ok else "fuzzy",
                    })
                    break

    out_csv = ANALYSIS_DIR / "sites_cleaned.csv"
    out_df["epsg_source"] = pd.to_numeric(out_df["epsg_source"], errors="coerce").astype("Int64")
    out_df.to_csv(out_csv, index=False, encoding="utf-8")

    report_path = ANALYSIS_DIR / "quality_report.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(f"OK: {len(df)} lignes → {len(out_df)} sites → {out_csv}")
    print(f"Rapport: {report_path}")


if __name__ == "__main__":
    main()
