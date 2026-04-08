"""Patriarche XLSX extractor with multi-strategy EA_IDENT parsing."""

from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import load_workbook

from src.domain.models.raw_record import RawRecord

logger = logging.getLogger(__name__)

_DATATION_KEYWORDS = re.compile(
    r"(?i)\b(?:age|âge|hallstatt|fer|bronze|gallo|romain|tène|la\s+tène|"
    r"néolith|mérovingien|médiéval|antiq|moderne|paléolith|mésol|"
    r"indéterminé|EURFER|EURBRO)\b"
)

_TYPE_KEYWORDS = re.compile(
    r"(?i)\b(?:habitat|tumulus|nécropole|sépulture|fosse|silo|four|forge|"
    r"tombe|villa|oppidum|fortification|dépôt|enclos|ferme|atelier|"
    r"grenier|fond de cabane|voie|fossé|fanum|sanctuaire|occupation|"
    r"inhumation|crémation|enceinte)\b"
)


_CHRONO_DBF_FIELDS = ("CHRONO", "EUR_PERIODE", "PERIODE")
_EUR_DECODE_MAP = {
    "EURFER": "Âge du Fer",
    "EURBRO": "Âge du Bronze",
    "EURGAL": "Gallo-romain",
    "EURNEO": "Néolithique",
    "EURMER": "Mérovingien",
    "EURMED": "Médiéval",
    "EURMOD": "Moderne",
    "EURPAL": "Paléolithique",
    "EURMES": "Mésolithique",
}


class PatriarcheExtractor:
    """Extract Patriarche XLSX into RawRecords with EA_IDENT parsing."""

    def __init__(self, *, dbf_path: Path | None = None) -> None:
        self._coords: dict[str, tuple[float, float]] = {}
        self._chrono_dbf: dict[str, str] = {}
        if dbf_path and dbf_path.exists():
            self._load_dbf_data(dbf_path)

    def _load_dbf_data(self, dbf_path: Path) -> None:
        from dbfread import DBF

        db = DBF(str(dbf_path), encoding="cp1252")
        for rec in db:
            code = str(rec.get("EA_NATCODE", "")).strip()
            x = rec.get("X_DEGRE")
            y = rec.get("Y_DEGRE")
            if code and x and y:
                try:
                    self._coords[code] = (float(x), float(y))
                except (TypeError, ValueError):
                    pass

            for field in _CHRONO_DBF_FIELDS:
                raw = str(rec.get(field, "")).strip()
                if raw and code:
                    decoded = _EUR_DECODE_MAP.get(raw.upper(), raw)
                    self._chrono_dbf[code] = decoded
                    break

        logger.info("Loaded %d coordinates, %d chrono from %s",
                     len(self._coords), len(self._chrono_dbf), dbf_path.name)

    def supported_formats(self) -> list[str]:
        return [".xlsx"]

    def extract(self, source_path: Path) -> list[RawRecord]:
        wb = load_workbook(source_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return []

        header = [str(c or "").strip() for c in rows[0]]
        col_idx = {name: i for i, name in enumerate(header)}

        records: list[RawRecord] = []
        for row in rows[1:]:
            if row is None or all(v is None for v in row):
                continue
            rec = self._parse_row(row, col_idx, str(source_path.resolve()))
            if rec:
                records.append(rec)

        logger.info("Patriarche %s: %d records extracted", source_path.name, len(records))
        return records

    def _parse_row(
        self, row: tuple, col_idx: dict[str, int], path_str: str
    ) -> RawRecord | None:
        def get(col: str) -> str:
            idx = col_idx.get(col)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        ea_ident = get("Identification_de_l_EA")
        ea_numero = get("Numero_de_l_EA")
        code_national = get("Code_national_de_l_EA")
        commune_col = get("Nom_de_la_commune")

        parts = [p.strip() for p in ea_ident.split(" / ")]
        commune, lieu_dit, type_mention, periode_mention = self._parse_parts(parts)

        if not commune and commune_col:
            commune = commune_col

        lat, lon = None, None
        lookup_key = code_national or ea_numero
        if code_national in self._coords:
            lon, lat = self._coords[code_national]
        elif ea_numero in self._coords:
            lon, lat = self._coords[ea_numero]

        chrono_dbf = self._chrono_dbf.get(code_national) or self._chrono_dbf.get(ea_numero)
        if not periode_mention and chrono_dbf:
            periode_mention = chrono_dbf

        extra = {
            "patriarche_ea": ea_numero or None,
            "patriarche_code_national": code_national or None,
        }
        if lieu_dit:
            extra["lieu_dit"] = lieu_dit
        if chrono_dbf:
            extra["chrono_dbf"] = chrono_dbf

        return RawRecord(
            raw_text=ea_ident,
            commune=commune,
            type_mention=type_mention,
            periode_mention=periode_mention,
            latitude_raw=lat,
            longitude_raw=lon,
            source_path=path_str,
            extraction_method="patriarche",
            extra=extra,
        )

    def _parse_parts(self, parts: list[str]) -> tuple[str | None, str | None, str | None, str | None]:
        """Parse EA_IDENT parts with heuristic type/datation detection.

        Expected format: id / ea_code / commune / (empty) / lieu_dit / ... (type/datation in any order)
        """
        if len(parts) < 3:
            return None, None, None, None

        commune = parts[2] if len(parts) > 2 else None
        lieu_dit = parts[4] if len(parts) > 4 and parts[4] else None

        type_mention: str | None = None
        periode_mention: str | None = None

        tail = parts[5:] if len(parts) > 5 else []
        for segment in tail:
            if not segment:
                continue
            if self._is_datation(segment) and not periode_mention:
                periode_mention = segment
            elif self._is_type(segment) and not type_mention:
                type_mention = segment
            elif not periode_mention and not type_mention:
                if _DATATION_KEYWORDS.search(segment):
                    periode_mention = segment
                else:
                    type_mention = segment

        return commune, lieu_dit, type_mention, periode_mention

    @staticmethod
    def _is_datation(text: str) -> bool:
        return bool(_DATATION_KEYWORDS.search(text))

    @staticmethod
    def _is_type(text: str) -> bool:
        return bool(_TYPE_KEYWORDS.search(text))
