"""AFEAF funerary dataset extractor with 2-level header reconstruction."""

from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import load_workbook

from src.domain.models.raw_record import RawRecord

logger = logging.getLogger(__name__)

_SITE_RE = re.compile(r"^([A-Za-zÀ-ÿ\-]+(?:\s+[A-Za-zÀ-ÿ\-]+)?)\s+(.+)$")


class AFEAFExtractor:
    """Extract AFEAF funerary dataset with hierarchical header reconstruction."""

    def __init__(self, *, sheet_name: str = "PF-hallstatt") -> None:
        self._sheet_name = sheet_name

    def supported_formats(self) -> list[str]:
        return [".xlsx"]

    def extract(self, source_path: Path) -> list[RawRecord]:
        wb = load_workbook(source_path, read_only=True, data_only=True)
        if self._sheet_name not in wb.sheetnames:
            ws = wb.active
        else:
            ws = wb[self._sheet_name]

        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(all_rows) < 3:
            return []

        columns = self._reconstruct_headers(all_rows[0], all_rows[1])
        records: list[RawRecord] = []

        for row in all_rows[2:]:
            if row is None or all(v is None or str(v).strip() in ("", "*") for v in row):
                continue
            row_dict = {columns[j]: row[j] for j in range(min(len(columns), len(row)))}
            rec = self._row_to_record(row_dict, str(source_path.resolve()))
            if rec:
                records.append(rec)

        logger.info("AFEAF %s: %d records extracted", source_path.name, len(records))
        return records

    @staticmethod
    def _reconstruct_headers(row0: tuple, row1: tuple) -> list[str]:
        """Build flat column names from 2-level header: 'group.sub' or 'group'."""
        columns: list[str] = []
        current_group = ""

        for j in range(max(len(row0), len(row1))):
            group = str(row0[j]).strip() if j < len(row0) and row0[j] is not None else ""
            sub = str(row1[j]).strip() if j < len(row1) and row1[j] is not None else ""

            if group:
                current_group = group

            if sub:
                col_name = f"{current_group}.{sub}" if current_group else sub
            elif current_group:
                col_name = current_group
            else:
                col_name = f"col_{j}"

            columns.append(col_name)

        return columns

    def _row_to_record(self, row: dict, path_str: str) -> RawRecord | None:
        dpt = str(row.get("info SITE.DPT") or "").strip()
        site_raw = str(row.get("info SITE.SITE") or "").strip()

        if not site_raw:
            return None

        commune, lieu_dit = self._parse_site(site_raw)

        funeraire = {}
        for key, val in row.items():
            if val is None or str(val).strip() in ("", "*"):
                continue
            funeraire[key] = val

        extra: dict = {
            "departement": dpt or None,
            "funeraire": funeraire,
        }
        if lieu_dit:
            extra["lieu_dit"] = lieu_dit

        datation_cols = {k: v for k, v in row.items() if "DATATION" in k.upper() and v and str(v).strip() not in ("", "*")}
        datation_str = " / ".join(str(v) for v in datation_cols.values()) if datation_cols else None

        return RawRecord(
            raw_text=f"{dpt} {site_raw}",
            commune=commune,
            type_mention="nécropole",
            periode_mention=datation_str,
            latitude_raw=None,
            longitude_raw=None,
            source_path=path_str,
            extraction_method="afeaf",
            extra=extra,
        )

    @staticmethod
    def _parse_site(site_raw: str) -> tuple[str, str | None]:
        """Split 'Colmar rue des Aunes' → ('Colmar', 'rue des Aunes')."""
        m = _SITE_RE.match(site_raw)
        if m:
            return m.group(1), m.group(2)
        return site_raw, None
