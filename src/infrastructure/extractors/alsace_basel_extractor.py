"""Alsace-Basel multi-sheet XLSX extractor with FK joins and thésaurus."""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd

from src.domain.models.raw_record import RawRecord

logger = logging.getLogger(__name__)

_TYPE_THESAURUS: dict[str, str] = {
    "hab_ouv": "habitat",
    "habitat ouvert": "habitat",
    "habitat": "habitat",
    "habitat groupé": "habitat",
    "nécropole": "nécropole",
    "necropole": "nécropole",
    "tumulus": "tumulus",
    "tombe": "nécropole",
    "sépulture": "nécropole",
    "oppidum": "oppidum",
    "fortification": "oppidum",
    "enceinte": "oppidum",
    "dépôt": "dépôt",
    "atelier": "atelier",
    "sanctuaire": "sanctuaire",
    "voie": "voie",
}


class AlsaceBaselExtractor:
    """Extract from multi-sheet Alsace-Basel XLSX with FK joins."""

    def supported_formats(self) -> list[str]:
        return [".xlsx"]

    def extract(self, source_path: Path) -> list[RawRecord]:
        sheets = self._read_sheets(source_path)
        if sheets is None:
            return []

        df_sites, df_occ, df_mob, thesaurus = sheets
        occ_by_site = df_occ.groupby("fk_site") if "fk_site" in df_occ.columns else None
        mob_by_occ = df_mob.groupby("fk_occupation") if "fk_occupation" in df_mob.columns else None

        records: list[RawRecord] = []
        for _, site in df_sites.iterrows():
            site_id = site.get("id_site")
            occupations = self._get_occupations(site_id, occ_by_site, mob_by_occ, thesaurus)

            type_mention = self._resolve_site_type(occupations)
            periode_mention = self._resolve_datation(occupations)

            raw_x = self._to_float(site.get("x"))
            raw_y = self._to_float(site.get("y"))
            epsg = self._to_int(site.get("epsg_coord")) or 4326

            lat, lon = None, None
            extra = {
                "id_site": site_id,
                "pays": str(site.get("pays") or "").strip(),
                "admin1": str(site.get("admin1") or "").strip(),
                "occupations": occupations,
            }

            if raw_x is not None and raw_y is not None:
                if epsg == 4326:
                    lon, lat = raw_x, raw_y
                elif epsg == 2154:
                    extra["x_l93"] = raw_x
                    extra["y_l93"] = raw_y
                    extra["epsg_source"] = 2154
                else:
                    from src.infrastructure.geocoding.reprojector import Reprojector
                    reproj = Reprojector()
                    x_l93, y_l93, in_bounds = reproj.to_lambert93(raw_x, raw_y, epsg)
                    if in_bounds:
                        extra["x_l93"] = x_l93
                        extra["y_l93"] = y_l93
                    extra["epsg_source"] = epsg

            commune = str(site.get("commune") or "").strip()
            lieu_dit = str(site.get("lieu_dit") or "").strip()
            if lieu_dit:
                extra["lieu_dit"] = lieu_dit

            records.append(RawRecord(
                raw_text=f"site {site_id}: {commune} - {lieu_dit}",
                commune=commune or None,
                type_mention=type_mention,
                periode_mention=periode_mention,
                latitude_raw=lat,
                longitude_raw=lon,
                source_path=str(source_path.resolve()),
                extraction_method="alsace_basel",
                extra=extra,
            ))

        logger.info("Alsace-Basel %s: %d site records", source_path.name, len(records))
        return records

    def _read_sheets(self, source_path: Path) -> tuple | None:
        try:
            sheets = pd.read_excel(source_path, sheet_name=None, engine="openpyxl")
        except (TypeError, Exception) as exc:
            logger.warning("pandas/openpyxl failed for %s: %s — trying manual repair", source_path.name, exc)
            sheets = self._read_sheets_repaired(source_path)
            if sheets is None:
                return None
            return sheets

        if isinstance(sheets, tuple):
            return sheets

        sheet_names = list(sheets.keys())
        if len(sheet_names) < 3:
            logger.warning("Expected >=3 sheets, got %d", len(sheet_names))
            return None

        df_sites = sheets.get("sites", sheets[sheet_names[0]])
        df_occ = sheets.get("occupations", sheets[sheet_names[1]])
        df_mob = sheets.get("mobilier", sheets[sheet_names[2]])

        thesaurus: dict[str, str] = {}
        df_thes = sheets.get("thésaurus")
        if df_thes is not None:
            for _, row in df_thes.iterrows():
                try:
                    val = row.iloc[2] if len(row) > 2 else None
                    if val is not None:
                        val_str = str(val).strip().lower()
                        if val_str and val_str != "nan":
                            thesaurus[val_str] = val_str
                except Exception:
                    pass

        return df_sites, df_occ, df_mob, thesaurus

    def _read_sheets_repaired(self, source_path: Path) -> tuple | None:
        """Workaround for openpyxl MultiCellRange bug: patch before loading."""
        import openpyxl.worksheet.datavalidation as dv_mod

        original_init = dv_mod.DataValidation.__init__

        def _patched_init(self_dv, *args, **kwargs):
            sqref = kwargs.get("sqref") or (args[6] if len(args) > 6 else None)
            if sqref is not None and not isinstance(sqref, str):
                kwargs["sqref"] = str(sqref)
                if len(args) > 6:
                    args = list(args)
                    args[6] = str(sqref)
                    args = tuple(args)
            try:
                original_init(self_dv, *args, **kwargs)
            except TypeError:
                kwargs.pop("sqref", None)
                original_init(self_dv, *args, **kwargs)

        dv_mod.DataValidation.__init__ = _patched_init
        try:
            from openpyxl import load_workbook
            wb = load_workbook(source_path, data_only=True)
        except Exception as exc:
            logger.error("openpyxl patched load also failed: %s", exc)
            dv_mod.DataValidation.__init__ = original_init
            return None
        finally:
            dv_mod.DataValidation.__init__ = original_init

        dfs: dict[str, pd.DataFrame] = {}
        for ws in wb.worksheets:
            data = list(ws.iter_rows(values_only=True))
            if not data:
                continue
            header = [str(c or f"col_{i}") for i, c in enumerate(data[0])]
            dfs[ws.title] = pd.DataFrame(data[1:], columns=header)
        wb.close()

        if len(dfs) < 3:
            logger.warning("Repaired read: expected >=3 sheets, got %d", len(dfs))
            return None

        sheet_names = list(dfs.keys())
        df_sites = dfs.get("sites", dfs[sheet_names[0]])
        df_occ = dfs.get("occupations", dfs[sheet_names[1]])
        df_mob = dfs.get("mobilier", dfs[sheet_names[2]])

        thesaurus: dict[str, str] = {}
        df_thes = dfs.get("thésaurus")
        if df_thes is not None:
            for _, row in df_thes.iterrows():
                try:
                    val = row.iloc[2] if len(row) > 2 else None
                    if val is not None:
                        val_str = str(val).strip().lower()
                        if val_str and val_str != "none":
                            thesaurus[val_str] = val_str
                except Exception:
                    pass

        logger.info("Repaired read: %s → %d sheets", source_path.name, len(dfs))
        return df_sites, df_occ, df_mob, thesaurus

    def _get_occupations(
        self, site_id, occ_by_site, mob_by_occ, thesaurus: dict
    ) -> list[dict]:
        if occ_by_site is None or site_id is None:
            return []
        try:
            occ_rows = occ_by_site.get_group(site_id)
        except KeyError:
            return []

        occupations = []
        for _, occ in occ_rows.iterrows():
            occ_id = occ.get("id_occupation")
            occ_type = str(occ.get("type") or "").strip()
            datation = str(occ.get("datation") or "").strip()
            comment = str(occ.get("commentaire_occupation") or "").strip()

            mobilier = []
            if mob_by_occ is not None and occ_id is not None:
                try:
                    mob_rows = mob_by_occ.get_group(occ_id)
                    for _, mob in mob_rows.iterrows():
                        mobilier.append({
                            "type_mobilier": str(mob.get("type_mobilier") or "").strip(),
                            "NR": mob.get("NR_mobilier"),
                            "NMI": mob.get("NMI_mobilier"),
                        })
                except KeyError:
                    pass

            occupations.append({
                "type": occ_type,
                "datation": datation,
                "commentaire": comment,
                "mobilier": mobilier,
            })

        return occupations

    @staticmethod
    def _resolve_site_type(occupations: list[dict]) -> str:
        for occ in occupations:
            occ_type = occ.get("type", "").lower()
            mapped = _TYPE_THESAURUS.get(occ_type)
            if mapped:
                return mapped
            for key, val in _TYPE_THESAURUS.items():
                if key in occ_type:
                    return val
        return "indéterminé"

    @staticmethod
    def _resolve_datation(occupations: list[dict]) -> str | None:
        datations = [occ["datation"] for occ in occupations if occ.get("datation")]
        return " / ".join(datations) if datations else None

    @staticmethod
    def _to_float(val) -> float | None:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _to_int(val) -> int | None:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return None
